#!/usr/bin/env python3

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os


def process_classification_json(json_file_path, output_excel_path):
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    rows = []
    fn_count = 0
    fp_count = 0
    fp_files = set()

    for file_name, file_data in data.items():
        evaluations = file_data.get('evaluations', [])

        if not evaluations:
            rows.append({
                'file_name': file_name,
                'rule_name': '(нет сработавших сигнатур)',
                'status': 'FN'
            })
            fn_count += 1
        else:
            unique_rules = {}
            for eval_item in evaluations:
                rule_name = eval_item.get('rule', 'Unknown Rule')
                status = eval_item.get('status', 'Unknown')

                if status not in ['TP', 'FP']:
                    continue

                if rule_name not in unique_rules:
                    unique_rules[rule_name] = status

            first_row = True
            for rule_name, status in unique_rules.items():
                rows.append({
                    'file_name': file_name if first_row else '',
                    'rule_name': rule_name,
                    'status': status
                })
                first_row = False

                if status == 'FP':
                    fp_count += 1
                    fp_files.add(file_name)

    df = pd.DataFrame(rows)
    df.to_excel(output_excel_path, index=False, sheet_name='Classification Results')
    format_excel(output_excel_path)

    print(f"\n{'='*50}")
    print(f"Обработка завершена")
    print(f"Входной файл: {json_file_path}")
    print(f"Выходной файл: {output_excel_path}")
    print(f"{'='*50}")
    print(f"Статистика:")
    print(f"  - Всего файлов в JSON: {len(data)}")
    print(f"  - Количество FN (файлов без сработавших сигнатур): {fn_count}")
    print(f"  - Количество FP (уникальных ложных срабатываний): {fp_count}")
    if fp_files:
        print(f"  - Файлы с FP ({len(fp_files)}): {', '.join(sorted(fp_files))}")
    else:
        print(f"  - Файлы с FP: нет")
    print(f"{'='*50}\n")


def format_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 12

        header_font = Font(bold=True, size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[0].alignment = left_alignment
            row[0].border = thin_border
            row[1].alignment = left_alignment
            row[1].border = thin_border
            row[2].alignment = center_alignment
            row[2].border = thin_border

            if row[2].value == 'TP':
                row[2].font = Font(color='006100', bold=True)
            elif row[2].value == 'FP':
                row[2].font = Font(color='9C6500', bold=True)
            elif row[2].value == 'FN':
                row[2].font = Font(color='9C0006', bold=True)

        ws.freeze_panes = 'A2'
        wb.save(excel_path)
        print("Форматирование Excel успешно применено")
    except Exception as e:
        print(f"Предупреждение: не удалось применить форматирование - {e}")


def main():
    INPUT_JSON = r"C:\Users\vboxuser\classification_with_out_rules_second_dataset.json"
    OUTPUT_EXCEL = r"C:\Users\vboxuser\classification_results_with_our_rules_second_dataset.xlsx"

    if not os.path.exists(INPUT_JSON):
        print(f"Ошибка: файл '{INPUT_JSON}' не найден")
        return

    try:
        process_classification_json(INPUT_JSON, OUTPUT_EXCEL)

        try:
            import subprocess
            subprocess.Popen(['start', OUTPUT_EXCEL], shell=True)
            print("Открываю Excel-файл...")
        except:
            pass

    except json.JSONDecodeError:
        print("Ошибка: файл не является корректным JSON")
    except Exception as e:
        print(f"Произошла ошибка: {e}")


if __name__ == "__main__":
    main()
